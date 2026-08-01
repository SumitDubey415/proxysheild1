import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

import database

def generate_weekly_excel_report(output_path, branch_filter=None):
    """
    Generate an Excel attendance matrix grid for the last 7 days.
    Grid format: Student Name, UID, Roll No, Branch, Day 1 to Day 7 status (✔/✘/!).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Attendance Report"

    # Define color styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Arial", size=16, bold=True, color="0F172A")
    subtitle_font = Font(name="Arial", size=10, italic=True, color="64748B")

    present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green
    present_font = Font(name="Arial", size=11, bold=True, color="166534")

    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
    absent_font = Font(name="Arial", size=11, bold=True, color="991B1B")

    proxy_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Soft Amber
    proxy_font = Font(name="Arial", size=11, bold=True, color="92400E")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Banner
    ws.merge_cells("A1:K1")
    ws["A1"] = "SMART ANTI-PROXY ATTENDANCE SYSTEM - WEEKLY REPORT"
    ws["A1"].font = title_font

    ws.merge_cells("A2:K2")
    today = datetime.now().date()
    dates = [(today - timedelta(days=6 - i)) for i in range(7)]
    start_str = dates[0].strftime("%b %d, %Y")
    end_str = dates[-1].strftime("%b %d, %Y")
    ws["A2"] = f"Period: {start_str} to {end_str} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = subtitle_font

    # Headers
    headers = ["#", "Student Name", "UID", "Roll No", "Branch"] + [d.strftime("%a (%m/%d)") for d in dates] + ["Att. %"]
    ws.append([]) # Row 3 blank
    ws.append(headers) # Row 4 headers

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Fetch students
    students = database.get_all_students()
    if branch_filter:
        students = [s for s in students if s["branch"].lower() == branch_filter.lower()]

    row_idx = 5
    for idx, student in enumerate(students, start=1):
        uid = student["uid"]
        row_data = [idx, student["name"], uid, student["roll_no"], student["branch"]]
        
        present_count = 0
        total_days = 7

        date_statuses = []
        for d in dates:
            d_str = d.strftime("%Y-%m-%d")
            records = database.get_attendance_records(date_filter=d_str, uid_filter=uid)
            if records:
                st = records[0]["status"]
                if st in ("PRESENT", "MANUAL_OVERRIDE"):
                    date_statuses.append(("✔", "PRESENT"))
                    present_count += 1
                elif st == "PROXY_ALERT":
                    date_statuses.append(("⚠ PROXY", "PROXY_ALERT"))
                elif st == "FLAGGED_NO_FACE":
                    date_statuses.append(("⚠ NO FACE", "PROXY_ALERT"))
                else:
                    date_statuses.append(("✘", "ABSENT"))
            else:
                date_statuses.append(("-", "NONE"))

        percentage = round((present_count / total_days * 100), 1)

        ws.cell(row=row_idx, column=1, value=idx).alignment = center_align
        ws.cell(row=row_idx, column=2, value=student["name"]).alignment = left_align
        ws.cell(row=row_idx, column=3, value=uid).alignment = center_align
        ws.cell(row=row_idx, column=4, value=student["roll_no"]).alignment = center_align
        ws.cell(row=row_idx, column=5, value=student["branch"]).alignment = center_align

        for i, (symbol, st_type) in enumerate(date_statuses, start=6):
            cell = ws.cell(row=row_idx, column=i, value=symbol)
            cell.alignment = center_align
            if st_type == "PRESENT":
                cell.fill = present_fill
                cell.font = present_font
            elif st_type == "ABSENT":
                cell.fill = absent_fill
                cell.font = absent_font
            elif st_type == "PROXY_ALERT":
                cell.fill = proxy_fill
                cell.font = proxy_font

        pct_cell = ws.cell(row=row_idx, column=13, value=f"{percentage}%")
        pct_cell.alignment = center_align
        pct_cell.font = Font(bold=True, color="166534" if percentage >= 75 else "991B1B")

        for c in range(1, 14):
            ws.cell(row=row_idx, column=c).border = thin_border

        row_idx += 1

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
    return output_path

def generate_weekly_pdf_report(output_path, branch_filter=None):
    """
    Generate a printable PDF weekly attendance grid.
    """
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter),
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=15
    )

    elements = []
    elements.append(Paragraph("Smart Anti-Proxy Attendance System - Weekly Grid Report", title_style))

    today = datetime.now().date()
    dates = [(today - timedelta(days=6 - i)) for i in range(7)]
    start_str = dates[0].strftime("%b %d")
    end_str = dates[-1].strftime("%b %d, %Y")
    
    elements.append(Paragraph(f"Period: <b>{start_str} - {end_str}</b> | Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))

    # Grid Table Headers
    date_headers = [d.strftime("%a %m/%d") for d in dates]
    headers = ["Name", "UID", "Branch"] + date_headers + ["Att %"]

    table_data = [headers]

    students = database.get_all_students()
    if branch_filter:
        students = [s for s in students if s["branch"].lower() == branch_filter.lower()]

    for student in students:
        uid = student["uid"]
        row = [student["name"][:15], uid, student["branch"]]

        present_count = 0
        for d in dates:
            d_str = d.strftime("%Y-%m-%d")
            records = database.get_attendance_records(date_filter=d_str, uid_filter=uid)
            if records:
                st = records[0]["status"]
                if st in ("PRESENT", "MANUAL_OVERRIDE"):
                    row.append("✔")
                    present_count += 1
                elif st in ("PROXY_ALERT", "FLAGGED_NO_FACE"):
                    row.append("⚠ ALERT")
                else:
                    row.append("✘")
            else:
                row.append("-")

        pct = f"{round((present_count / 7 * 100), 1)}%"
        row.append(pct)
        table_data.append(row)

    col_widths = [110, 70, 70] + [65] * 7 + [55]

    pdf_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))

    elements.append(pdf_table)
    doc.build(elements)
    return output_path
